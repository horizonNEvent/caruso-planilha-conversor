import streamlit as st
import openpyxl
import io
import os

# Configuração da página
st.set_page_config(
    page_title="Caruso - Conversor Excel para TXT",
    page_icon="📄",
    layout="centered"
)

def formatar_valor(valor, tamanho):
    """
    Formata o valor para um tamanho fixo, preenchendo com espaços ou truncando.
    """
    s_valor = str(valor).strip() if valor is not None else ""
    if s_valor.lower() in ['nan', 'none', 'null']:
        s_valor = ''
    
    if len(s_valor) > tamanho:
        return s_valor[:tamanho]
    
    return s_valor.ljust(tamanho)

def processar_excel_streamlit(uploaded_file):
    """
    Processa o arquivo Excel enviado e retorna o conteúdo do TXT em memória.
    """
    try:
        # Carregar o arquivo Excel em memória
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        
        if 'preencher' not in wb.sheetnames:
            return None, "Erro: A aba 'preencher' não foi encontrada no arquivo Excel."
            
        ws = wb['preencher']

        # Obter comprimentos
        comprimentos = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                try:
                    comprimentos.append(int(val))
                except (ValueError, TypeError):
                    break
            else:
                break
        
        if not comprimentos:
            return None, "Erro: Nenhum comprimento de campo encontrado na linha 1."

        output = io.StringIO()
        processadas = 0

        # Processamento linha a linha
        for row_idx in range(3, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is None:
                continue
            
            # Parte 1: Colunas A até BF
            limite1 = min(58, len(comprimentos))
            linha1_partes = []
            for i in range(0, limite1):
                valor = ws.cell(row=row_idx, column=i+1).value
                tamanho = comprimentos[i]
                linha1_partes.append(formatar_valor(valor, tamanho))
            
            output.write("".join(linha1_partes) + '\n')

            # Parte 2: Colunas BG em diante
            if len(comprimentos) > 58:
                linha2_partes = []
                for i in range(58, len(comprimentos)):
                    valor = ws.cell(row=row_idx, column=i+1).value
                    tamanho = comprimentos[i]
                    linha2_partes.append(formatar_valor(valor, tamanho))
                
                output.write("".join(linha2_partes) + '\n')
            
            processadas += 1

        content = output.getvalue()
        output.close()
        return content, f"Sucesso! {processadas} linhas processadas."

    except Exception as e:
        return None, f"Erro inesperado: {str(e)}"

# Interface Streamlit
st.title("📄 Conversor de Planilha Caruso")
st.markdown("""
Esta ferramenta converte sua planilha Excel para o formato TXT de largura fixa, 
seguindo as regras da aba **'preencher'**.
""")

uploaded_file = st.file_uploader("Escolha o arquivo Excel (.xlsx)", type="xlsx")

if uploaded_file is not None:
    st.info(f"Arquivo carregado: **{uploaded_file.name}**")
    
    if st.button("Converter para TXT", type="primary"):
        with st.spinner("Processando..."):
            txt_content, message = processar_excel_streamlit(uploaded_file)
            
            if txt_content:
                st.success(message)
                
                # Definir nome do arquivo de saída
                base_name = os.path.splitext(uploaded_file.name)[0]
                output_filename = f"{base_name}_resultado.txt"
                
                # Botão de download
                st.download_button(
                    label="📥 Baixar Arquivo TXT",
                    data=txt_content.encode('utf-8'),
                    file_name=output_filename,
                    mime="text/plain"
                )
            else:
                st.error(message)

st.divider()
st.caption("Desenvolvido para automação de Planilhas Caruso.")
