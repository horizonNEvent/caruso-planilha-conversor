import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

def formatar_valor(valor, tamanho):
    """
    Formata o valor para um tamanho fixo, preenchendo com espaços ou truncando.
    Garante que valores numéricos tenham 2 casas decimais, mas REMOVE a vírgula/ponto para o TXT.
    Exemplo: 1000.00 vira "100000" (preenchido com espaços até o tamanho).
    """
    if valor is None:
        return "".ljust(tamanho)
    
    s_valor = str(valor).strip()
    if s_valor.lower() in ['nan', 'none', 'null', '']:
        return "".ljust(tamanho)

    # Tenta reconhecer se é um número
    try:
        if isinstance(valor, str):
            if any(c.isalpha() for c in s_valor.replace('e', '').replace('E', '')):
                raise ValueError
            test_valor = s_valor.replace('.', '').replace(',', '.')
            v_float = float(test_valor)
        else:
            v_float = float(valor)
        
        # Formata com 2 casas decimais e remove o ponto (ex: 100.00 -> "10000")
        s_valor = f"{v_float:.2f}".replace('.', '').replace(',', '')
    except (ValueError, TypeError):
        pass
    
    if len(s_valor) > tamanho:
        return s_valor[:tamanho]
    
    return s_valor.ljust(tamanho)

def processar_excel(caminho_arquivo_excel, caminho_arquivo_txt, log_callback=None):
    try:
        if log_callback: log_callback(f"Abrindo arquivo: {os.path.basename(caminho_arquivo_excel)}...")
        
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        
        # Verificar se a aba 'preencher' existe
        if 'preencher' not in wb.sheetnames:
            raise ValueError("A aba 'preencher' não foi encontrada no arquivo Excel.")
            
        ws = wb['preencher']

        # Obter os comprimentos dos campos da primeira linha (row 1)
        comprimentos = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                try:
                    comprimentos.append(int(val))
                except ValueError:
                    break
            else:
                break
        
        if not comprimentos:
            raise ValueError("Nenhum comprimento de campo encontrado na primeira linha.")

        total_linhas = ws.max_row - 2 # Linha 3 em diante
        processadas = 0

        with open(caminho_arquivo_txt, 'w', encoding='utf-8') as arquivo_txt:
            for row_idx in range(3, ws.max_row + 1):
                # Verificar se a linha está vazia (pelo menos a primeira coluna)
                if ws.cell(row=row_idx, column=1).value is None:
                    continue
                
                # Linha 1: Colunas A até BF (índices 0 a 57, colunas 1 a 58)
                # Garantir que não tentamos acessar índices fora de 'comprimentos'
                limite1 = min(58, len(comprimentos))
                linha1_partes = []
                for i in range(0, limite1):
                    valor = ws.cell(row=row_idx, column=i+1).value
                    tamanho = comprimentos[i]
                    linha1_partes.append(formatar_valor(valor, tamanho))
                
                arquivo_txt.write("".join(linha1_partes) + '\n')

                # Linha 2: Colunas BG até o final (índices 58 a ...)
                if len(comprimentos) > 58:
                    linha2_partes = []
                    for i in range(58, len(comprimentos)):
                        valor = ws.cell(row=row_idx, column=i+1).value
                        tamanho = comprimentos[i]
                        linha2_partes.append(formatar_valor(valor, tamanho))
                    
                    arquivo_txt.write("".join(linha2_partes) + '\n')
                
                processadas += 1

        if log_callback: log_callback(f"Sucesso! {processadas} linhas processadas.")
        return True, f"Arquivo '{os.path.basename(caminho_arquivo_txt)}' gerado com sucesso!"
    except Exception as e:
        if log_callback: log_callback(f"Erro: {str(e)}")
        return False, str(e)

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Conversor Excel para TXT")
        self.root.geometry("600x400")
        self.root.configure(bg="#f0f0f0")

        self.caminho_excel = tk.StringVar()
        
        # Estilo
        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", background="#0078d7")
        style.configure("Main.TFrame", background="#f0f0f0")

        main_frame = ttk.Frame(root, padding="20", style="Main.TFrame")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Título
        title_label = tk.Label(main_frame, text="Conversor de Planilha Caruso", font=("Segoe UI", 16, "bold"), bg="#f0f0f0", fg="#333")
        title_label.pack(pady=(0, 20))

        # Seleção de arquivo
        file_frame = ttk.Frame(main_frame, style="Main.TFrame")
        file_frame.pack(fill=tk.X, pady=10)

        entry_file = ttk.Entry(file_frame, textvariable=self.caminho_excel, width=50)
        entry_file.pack(side=tk.LEFT, padx=(0, 10), expand=True, fill=tk.X)

        btn_browse = ttk.Button(file_frame, text="Procurar...", command=self.browse_file)
        btn_browse.pack(side=tk.RIGHT)

        # Botão Processar
        self.btn_convert = ttk.Button(main_frame, text="Converter para TXT", command=self.start_conversion)
        self.btn_convert.pack(pady=20)

        # Área de Log
        log_label = tk.Label(main_frame, text="Status:", bg="#f0f0f0", font=("Segoe UI", 10))
        log_label.pack(anchor=tk.W)

        self.log_text = tk.Text(main_frame, height=8, state=tk.DISABLED, bg="white", font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def browse_file(self):
        filename = filedialog.askopenfilename(
            title="Selecione a planilha Excel",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if filename:
            self.caminho_excel.set(filename)
            self.log("Arquivo selecionado: " + os.path.basename(filename))

    def log(self, message):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def start_conversion(self):
        caminho = self.caminho_excel.get()
        if not caminho:
            messagebox.showwarning("Aviso", "Por favor, selecione um arquivo Excel primeiro.")
            return

        if not os.path.exists(caminho):
            messagebox.showerror("Erro", "O arquivo selecionado não existe.")
            return

        nome_base = os.path.splitext(caminho)[0]
        caminho_txt = f"{nome_base}_resultado.txt"

        self.btn_convert.config(state=tk.DISABLED)
        self.log("Iniciando processamento...")
        
        success, msg = processar_excel(caminho, caminho_txt, self.log)
        
        if success:
            messagebox.showinfo("Sucesso", msg)
        else:
            messagebox.showerror("Erro", msg)
            
        self.btn_convert.config(state=tk.NORMAL)

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
