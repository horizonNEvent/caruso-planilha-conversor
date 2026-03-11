import openpyxl
import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

def formatar_valor(valor, tamanho):
    """
    Formata o valor para um tamanho fixo, preenchendo com espaços ou truncando.
    Garante que valores numéricos tenham 2 casas decimais, mas REMOVE a vírgula/ponto para o TXT.
    Exemplo: 1000.00 vira "100000" (preenchido com espaços até o tamanho).
    """
    if valor is None:
        return "".ljust(tamanho)
    
    s_valor = str(valor).strip()
    if s_valor.lower() in ['nan', 'none', 'null', '']:
        return "".ljust(tamanho)

    # Tenta reconhecer se é um número
    try:
        if isinstance(valor, str):
            if any(c.isalpha() for c in s_valor.replace('e', '').replace('E', '')):
                raise ValueError
            test_valor = s_valor.replace('.', '').replace(',', '.')
            v_float = float(test_valor)
        else:
            v_float = float(valor)
        
        # Formata com 2 casas decimais e remove o ponto (ex: 100.00 -> "10000")
        s_valor = f"{v_float:.2f}".replace('.', '').replace(',', '')
    except (ValueError, TypeError):
        pass
    
    if len(s_valor) > tamanho:
        return s_valor[:tamanho]
    
    return s_valor.ljust(tamanho)

def processar_excel(caminho_arquivo_excel, caminho_arquivo_txt, log_callback=None):
    """
    Lógica de processamento do arquivo Excel para TXT seguindo as regras da aba 'preencher'.
    """
    try:
        if log_callback: log_callback(f"Abrindo arquivo: {os.path.basename(caminho_arquivo_excel)}...")
        
        # Carregar o arquivo Excel
        wb = openpyxl.load_workbook(caminho_arquivo_excel, data_only=True)
        
        # Verificar se a aba 'preencher' existe
        if 'preencher' not in wb.sheetnames:
            error_msg = "Erro: A aba 'preencher' não foi encontrada no arquivo Excel."
            if log_callback: log_callback(error_msg)
            return False, error_msg
            
        ws = wb['preencher']

        # Obter os comprimentos dos campos da primeira linha (row 1)
        comprimentos = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                try:
                    comprimentos.append(int(val))
                except (ValueError, TypeError):
                    break
            else:
                break
        
        if not comprimentos:
            error_msg = "Erro: Nenhum comprimento de campo encontrado na linha 1."
            if log_callback: log_callback(error_msg)
            return False, error_msg

        processadas = 0
        total_rows = ws.max_row

        with open(caminho_arquivo_txt, 'w', encoding='utf-8') as arquivo_txt:
            # O processamento começa na linha 3 (row 3)
            for row_idx in range(3, total_rows + 1):
                # Verificar se a linha está vazia (pelo menos a primeira coluna)
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val is None:
                    continue
                
                # Parte 1: Colunas A até BF (índices 0 a 57, colunas 1 a 58)
                limite1 = min(58, len(comprimentos))
                linha1_partes = []
                for i in range(0, limite1):
                    valor = ws.cell(row=row_idx, column=i+1).value
                    tamanho = comprimentos[i]
                    linha1_partes.append(formatar_valor(valor, tamanho))
                
                arquivo_txt.write("".join(linha1_partes) + '\n')

                # Parte 2: Colunas BG até o final (índices 58 em diante)
                if len(comprimentos) > 58:
                    linha2_partes = []
                    for i in range(58, len(comprimentos)):
                        valor = ws.cell(row=row_idx, column=i+1).value
                        tamanho = comprimentos[i]
                        linha2_partes.append(formatar_valor(valor, tamanho))
                    
                    arquivo_txt.write("".join(linha2_partes) + '\n')
                
                processadas += 1

        success_msg = f"Sucesso! {processadas} linhas processadas e salvas em {os.path.basename(caminho_arquivo_txt)}."
        if log_callback: log_callback(success_msg)
        return True, success_msg

    except Exception as e:
        error_msg = f"Erro inesperado: {str(e)}"
        if log_callback: log_callback(error_msg)
        return False, error_msg

class ConversorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Caruso - Conversor Excel para TXT")
        self.root.geometry("650x450")
        self.root.minsize(500, 400)
        
        # Tema e Estilo
        self.root.configure(bg="#f5f6f7")
        style = ttk.Style()
        style.theme_use('clam')
        
        # Frame Principal
        main_frame = ttk.Frame(root, padding="30")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Cabeçalho
        header_label = tk.Label(
            main_frame, 
            text="Conversor de Planilha", 
            font=("Segoe UI", 18, "bold"), 
            bg="#f5f6f7", 
            fg="#2c3e50"
        )
        header_label.pack(pady=(0, 20))

        # Seção de Seleção de Arquivo
        file_container = ttk.LabelFrame(main_frame, text=" Selecione o arquivo Excel (.xlsx) ", padding="15")
        file_container.pack(fill=tk.X, pady=10)

        self.path_var = tk.StringVar()
        entry_path = ttk.Entry(file_container, textvariable=self.path_var)
        entry_path.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        btn_search = ttk.Button(file_container, text="Procurar...", command=self.search_file)
        btn_search.pack(side=tk.RIGHT)

        # Botão de Ação
        self.btn_convert = tk.Button(
            main_frame, 
            text="INICIAR CONVERSÃO", 
            command=self.run_conversion,
            bg="#27ae60", 
            fg="white", 
            font=("Segoe UI", 11, "bold"),
            relief="flat",
            padx=20,
            pady=10,
            cursor="hand2",
            activebackground="#219150",
            activeforeground="white"
        )
        self.btn_convert.pack(pady=25)

        # Área de Logs/Status
        status_frame = ttk.LabelFrame(main_frame, text=" Logs de Operação ", padding="10")
        status_frame.pack(fill=tk.BOTH, expand=True)

        self.log_area = tk.Text(
            status_frame, 
            height=8, 
            state=tk.DISABLED, 
            font=("Consolas", 9),
            bg="#ffffff",
            fg="#34495e",
            borderwidth=1,
            relief="solid"
        )
        self.log_area.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar para o log
        scrollbar = ttk.Scrollbar(self.log_area, command=self.log_area.yview)
        self.log_area.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def search_file(self):
        filename = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        )
        if filename:
            self.path_var.set(os.path.normpath(filename))
            self.log(f"Arquivo selecionado: {os.path.basename(filename)}")

    def log(self, text):
        self.log_area.config(state=tk.NORMAL)
        self.log_area.insert(tk.END, f"[{os.path.basename(sys.argv[0])}] {text}\n")
        self.log_area.see(tk.END)
        self.log_area.config(state=tk.DISABLED)
        self.root.update_idletasks()

    def run_conversion(self):
        input_path = self.path_var.get()
        if not input_path:
            messagebox.showwarning("Aviso", "Por favor, selecione um arquivo Excel (.xlsx) primeiro.")
            return

        if not os.path.exists(input_path):
            messagebox.showerror("Erro", "O arquivo especificado não foi encontrado.")
            return

        # Definir nome de saída
        dir_name = os.path.dirname(input_path)
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(dir_name, f"{base_name}_resultado.txt")

        # Bloquear botão enquanto processa
        self.btn_convert.config(state=tk.DISABLED, bg="#95a5a6", cursor="arrow")
        self.log("Limpando ambiente e iniciando processamento...")

        success, message = processar_excel(input_path, output_path, self.log)

        if success:
            messagebox.showinfo("Sucesso", "Conversão concluída com sucesso!")
        else:
            messagebox.showerror("Erro", message)

        # Reativar botão
        self.btn_convert.config(state=tk.NORMAL, bg="#27ae60", cursor="hand2")

if __name__ == "__main__":
    root = tk.Tk()
    app = ConversorApp(root)
    # Se houver um arquivo no diretório atual, selecioná-lo automaticamente como sugestão
    excel_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
    if excel_files:
        app.path_var.set(os.path.abspath(excel_files[0]))
    
    root.mainloop()

